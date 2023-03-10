from django.shortcuts import render
from django.http import HttpResponse
from .forms import Insert_automation_data
from django.views import generic
from .models import automation_db
import openpyxl

# Create your views here.
def home(request):
    return render(request, "Index.html", context={})

class index(generic.ListView):
    model = automation_db
    template_name = 'Index.html'

    """ def get_context_data(self, *, object_list=None, **kwargs):
        context = (automation_db,self).get_context_data(**kwargs)
        context['reg_scripts'] = 
    """

class Insert_automation_data_view(generic.CreateView):
    form_class = Insert_automation_data
    model = automation_db
    template_name = 'Insert_automation_data.html'
    success_url = '/list/createdata'

class table_automation_view(generic.ListView):
    model = automation_db
    template_name = 'table_automation.html'

def upload_excel(request):
    if "GET" == request.method:
        return render(request, 'dashboard_app/Index.html', {})
    else:
        excel_file = request.FILES["excel_data"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1:A10"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        model = automation_db

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        #return render(request, 'dashboard_app/Index.html', {"excel_data": excel_data})
